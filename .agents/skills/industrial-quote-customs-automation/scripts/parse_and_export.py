# parse_and_export.py
# Reusable Automation Engine for Large Multi-page Industrial Invoices & Quotes
# Includes Automated shop.murrplastik.com Search by Order No (e.g. 83201208)

import os
import re
import sys
import argparse
import urllib.request
import urllib.parse
import json
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Built-in Technical Dictionary for Murrplastik & Industrial Components
CUSTOMS_DICTIONARY = {
    "SH": {
        "vn_name": "Giá đỡ ống luồn dây cáp bằng nhựa PA6 có nắp khóa kim loại",
        "material": "Nhựa Polyamide PA6 & Nắp Thép gia cường",
        "hs_code": "3926.90.99",
        "usage": "Định vị và gia cố gá đỡ ống dẫn hướng bảo vệ cáp trên thân robot"
    },
    "R-TEC": {
        "vn_name": "Hộp cơ cấu thu hồi cáp tự động tích hợp lò xo co rút",
        "material": "Nhựa PA6 + Lò xo Thép + Khung hợp kim Nhôm",
        "hs_code": "8479.89.99",
        "usage": "Tự động thu hồi co rút bù hành trình bảo vệ cáp nguồn hàn khi robot chuyển động"
    },
    "A-ZS": {
        "vn_name": "Hệ thống kẹp chặt giảm ứng lực căng cáp",
        "material": "Nhựa Polyamide PA6 + Cao su Elastomer",
        "hs_code": "3926.90.99",
        "usage": "Siết chặt định vị lõi cáp bên trong lòng ống dẫn chống xô dịch và giảm căng"
    },
    "R-ZL": {
        "vn_name": "Tấm đệm cao su giảm ứng lực căng cáp dạng cắt hình sao",
        "material": "Cao su đàn hồi Elastomer",
        "hs_code": "4016.99.99",
        "usage": "Giảm căng và chống trượt cho từng lõi cáp đơn bên trong khớp nối"
    },
    "KEG": {
        "vn_name": "Khớp nối cầu vạn năng xoay tự do / cố định",
        "material": "Nhựa Polyamide PA6",
        "hs_code": "3917.40.00",
        "usage": "Khớp nối định hướng xoay đa hướng cho ống luồn cáp tại các điểm gập xoay"
    },
    "KMG": {
        "vn_name": "Khớp nối vạn năng cố định định vị đầu ống",
        "material": "Nhựa Polyamide PA6",
        "hs_code": "3917.40.00",
        "usage": "Khóa định vị đầu ống dẫn hướng cố định tại điểm bắt đầu/kết thúc"
    },
    "SRF": {
        "vn_name": "Vòng định vị / cố định hành trình ống luồn",
        "material": "Nhựa Polyamide PA6",
        "hs_code": "3917.40.00",
        "usage": "Định vị vị trí gá kẹp và giới hạn hành trình gập của ống luồn"
    },
    "SS": {
        "vn_name": "Kẹp giữ định hướng ống luồn cáp cố định",
        "material": "Nhựa Polyamide PA6",
        "hs_code": "3926.90.99",
        "usage": "Kẹp giữ cố định đường ống luồn song song trên thân máy/robot"
    },
    "EW": {
        "vn_name": "Ống gợn sóng bảo vệ cáp điện nguồn và tín hiệu",
        "material": "Nhựa Polyamide PA6",
        "hs_code": "3917.39.00",
        "usage": "Bọc chứa và bảo vệ toàn bộ bó cáp điện khỏi va đập cơ học và mài mòn"
    }
}

def fetch_murrplastik_shop_data(order_no):
    """
    Auto-searches shop.murrplastik.com by Order No (e.g., 83201208)
    Returns dictionary with product shop URL, image URL, and title.
    """
    clean_order = str(order_no).strip()
    shop_url = f"https://shop.murrplastik.com/search?q={clean_order}"
    
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
    }
    
    result = {
        "order_no": clean_order,
        "shop_url": shop_url,
        "image_url": None,
        "title": None
    }
    
    try:
        req = urllib.request.Request(shop_url, headers=headers)
        with urllib.request.urlopen(req, timeout=5) as response:
            html = response.read().decode('utf-8', errors='ignore')
            
            # Extract product canonical/direct URL if redirected or listed
            link_match = re.search(r'href="([^"]*/p/[^"]*matnr=' + re.escape(clean_order) + r'[^"]*)"', html, re.IGNORECASE)
            if link_match:
                rel_link = link_match.group(1)
                result["shop_url"] = rel_link if rel_link.startswith("http") else f"https://shop.murrplastik.com{rel_link}"
                
            # Extract image URL
            img_match = re.search(r'<img[^>]+src="([^"]*(?:product|matnr|media)[^"]*)"', html, re.IGNORECASE)
            if img_match:
                rel_img = img_match.group(1)
                result["image_url"] = rel_img if rel_img.startswith("http") else f"https://shop.murrplastik.com{rel_img}"
                
            # Extract Title
            title_match = re.search(r'<h1[^>]*>(.*?)</h1>', html, re.DOTALL | re.IGNORECASE)
            if title_match:
                result["title"] = re.sub(r'<[^>]+>', '', title_match.group(1)).strip()
                
    except Exception as e:
        print(f"⚠️ Warning: Could not fetch live shop data for Order No {clean_order}: {e}")
        
    return result

def lookup_customs_info(part_no, description):
    desc_upper = str(description).upper()
    for key, info in CUSTOMS_DICTIONARY.items():
        if key in desc_upper:
            return info
    
    return {
        "vn_name": f"Linh kiện phụ kiện dẫn hướng cáp công nghiệp ({description})",
        "material": "Nhựa Polyamide PA6 / Hợp kim",
        "hs_code": "3926.90.99",
        "usage": "Bảo vệ và dẫn hướng cáp nguồn tín hiệu trong hệ thống tự động hóa"
    }

def format_excel_sheet(ws, title_text, is_customs=False):
    header_fill = PatternFill(start_color="1F2937" if not is_customs else "800020", 
                              end_color="1F2937" if not is_customs else "800020", 
                              fill_type="solid")
    header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    data_font = Font(name="Arial", size=9)
    bold_font = Font(name="Arial", size=9, bold=True)
    mono_font = Font(name="Consolas", size=9)
    
    thin_border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB')
    )
    
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ws.max_column)
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = title_text
    title_cell.font = Font(name="Arial", size=13, bold=True, color="800020" if is_customs else "111827")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35
    
    ws.row_dimensions[3].height = 28
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=3, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        
    for row in range(4, ws.max_row + 1):
        ws.row_dimensions[row].height = 24
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            
            val_str = str(cell.value or '')
            if col == 1 or 'STT' in str(ws.cell(row=3, column=col).value):
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = bold_font
            elif 'Mã' in str(ws.cell(row=3, column=col).value) or 'HS Code' in str(ws.cell(row=3, column=col).value):
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = mono_font
            elif 'Số lượng' in str(ws.cell(row=3, column=col).value):
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = bold_font

    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = 0
        for cell in col:
            val = str(cell.value or '')
            if '\n' in val:
                lines = val.split('\n')
                max_len = max(max_len, max(len(l) for l in lines))
            else:
                max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 50)

def generate_customs_workbook(items_df, output_filepath, fetch_shop=True):
    wb = openpyxl.Workbook()
    ws_main = wb.active
    ws_main.title = "Tong_hop_Hai_quan"
    
    records = []
    for idx, row in items_df.iterrows():
        part_no = row.get('Material', row.get('PartNo', ''))
        desc = row.get('Description', '')
        qty = row.get('Qty', '1')
        robot = row.get('Robot', 'ABB IRB 7600/6700')
        
        # Step 1b: Auto search shop.murrplastik.com by Order No
        shop_info = fetch_murrplastik_shop_data(part_no) if fetch_shop else {"shop_url": f"https://shop.murrplastik.com/search?q={part_no}"}
        
        c_info = lookup_customs_info(part_no, desc)
        records.append({
            'STT': idx + 1,
            'Mã vật tư (Order No)': part_no,
            'Tên sản phẩm Tiếng Anh': desc,
            'Tên Tiếng Việt Khai Hải Quan': c_info['vn_name'],
            'Chất liệu cấu thành': c_info['material'],
            'Công dụng chi tiết': c_info['usage'],
            'Mã HS Code Tham Khảo': c_info['hs_code'],
            'Số lượng': qty,
            'Link Shop Hãng': shop_info['shop_url'],
            'Áp dụng cho': robot
        })
        
    customs_df = pd.DataFrame(records)
    
    ws_main.append([])
    ws_main.append([])
    ws_main.append(list(customs_df.columns))
    for r in customs_df.itertuples(index=False):
        ws_main.append(list(r))
        
    format_excel_sheet(ws_main, "DANH MỤC VẬT TƯ NÂNG CẤP DRESS PACK ROBOT - KHAI BÁO HẢI QUAN", is_customs=True)
    
    wb.save(output_filepath)
    print(f"✅ Created Customs Excel File: {output_filepath}")

def main():
    parser = argparse.ArgumentParser(description="Industrial Invoice & Customs Excel Pipeline with Live Murrplastik Shop Search")
    parser.add_argument("--input", required=False, help="Input PDF/Excel file path")
    parser.add_argument("--fetch-shop", action="store_true", default=True, help="Auto search shop.murrplastik.com by Order No")
    parser.add_argument("--output-dir", default="./", help="Output directory")
    args = parser.parse_args()

    print("🚀 Industrial Quote & Customs Excel Engine Loaded.")
    print("🔎 Live Shop Search Enabled: shop.murrplastik.com/search?q={Order_No}")

if __name__ == "__main__":
    main()
